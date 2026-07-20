"""Excel price-list parser (openpyxl)."""

from __future__ import annotations

from abc import ABC, abstractmethod
from collections.abc import Iterator
from typing import Any

import openpyxl

from backend.utils.logging import get_logger

logger = get_logger(__name__)


class PriceDataSource(ABC):
    TEMPLATES: dict[str, list[str]] = {
        "name": ["Наименование"],
        "meas": ["Единиц", "Ед. изм.", "Ед изм"],
        "price": ["Цена"],
        "date": ["Дата"],
        "group": ["Родитель", "Группа"],
    }

    def _check_col_header_datatype(self, col_header: str, data_type: str) -> bool:
        for template in self.TEMPLATES[data_type]:
            if template.casefold() in col_header.casefold():
                return True
        return False

    @abstractmethod
    def iter_rows(self) -> Iterator[dict[str, str]]:
        raise NotImplementedError


class PriceExcelSource(PriceDataSource):
    LINES_TO_SEARCH = 5

    def __init__(self, xls_file_name: str, work_sheet_name: str | None = None) -> None:
        self._file_name = xls_file_name
        self._work_sheet_name = work_sheet_name
        self._loaded = False
        self._work_book: Any = None
        self._work_sheet: Any = None
        self._cols: dict[str, int | None] = {}

    def close(self) -> None:
        if self._work_book is not None:
            self._work_book.close()
            self._work_book = None
            self._loaded = False

    def _load(self) -> None:
        self._work_book = openpyxl.load_workbook(self._file_name)
        sheet = self._work_book.active
        if self._work_sheet_name is not None:
            sheet = self._work_book[self._work_sheet_name]
        if sheet is None:
            raise OSError(f"reading data from {self._file_name} error.")
        self._work_sheet = sheet
        self._cols = dict.fromkeys(self.TEMPLATES, None)
        min_row = sheet.min_row
        max_row = min_row + self.LINES_TO_SEARCH
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=sheet.min_column,
            max_col=sheet.max_column,
        ):
            for cell in row:
                for data_type in self._cols:
                    if self._check_col_header_datatype(str(cell.value), data_type):
                        self._cols[data_type] = int(cell.col_idx)
                        break
        self._loaded = True

    def iter_rows(self) -> Iterator[dict[str, str]]:
        if not self._loaded:
            self._load()
        sheet = self._work_sheet
        for row in range(sheet.min_row, sheet.max_row + 1):
            res: dict[str, str] = {}
            for data_type, col in self._cols.items():
                if col is None:
                    res[data_type] = ""
                else:
                    res[data_type] = str(sheet.cell(row=row, column=col).value)
            yield res
